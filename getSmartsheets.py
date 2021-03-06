"""
Copyright (c) 2019

Permission is hereby granted, free of charge, to any person obtaining a copy of
this software and associated documentation files (the "Software"), to deal in
the Software without restriction, including without limitation the rights to
use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies
of the Software, and to permit persons to whom the Software is furnished to do
so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
import csv
import os
import platform
import sys

import colorama
import smartsheet
from openpyxl import load_workbook
from smartsheet import reports


#class Bcolors:
#    """color palette for print statements"""
#    HEADER = '\033[96m'
#    OKBLUE = '\033[94m'
#    OKGREEN = '\033[92m'
#    ENDC = '\033[0m'
#    BOLD = '\033[1m'
#    UNDERLINE = '\033[4m'
#    WARNING = '\033[93m'
#    FAIL = '\033[91m'


class Bcolors:
    """color palette for print statements"""
    HEADER = ''
    OKBLUE = ''
    OKGREEN = ''
    ENDC = ''
    BOLD = ''
    UNDERLINE = ''
    WARNING = '-- Warning -- '
    FAIL = '\n\n ***************************************************   Error  ***************************************************\n'


def get_file_info(filename):
    """go get the file information"""
    try:
        ss_hndl = open(filename, 'r')
    except FileNotFoundError:
        print(Bcolors.FAIL + f"File not found {filename}" + Bcolors.ENDC)
        return(0)
    except Exception as e:
        errno, strerror = e.args
        print(Bcolors.FAIL + f"I/O error({errno}): {strerror} {filename}" + Bcolors.ENDC)
        return(0)
    else:
        line_list = ss_hndl.readlines()
        # strip off newlines
        info_list = []
        for line in line_list:
            info_list.append(line.rstrip())
        ss_hndl.close()
        return info_list


def convert_downloaded_files_to_tab_delimited(program_input_path):
    """let's open each new downloaded file and search for filename"""
    try:
        file_hndl = open(program_input_path + "filenames.txt", 'w')
    except Exception:
        print(Bcolors.FAIL + "Cannot open file " + program_input_path + " filenames.txt" + Bcolors.ENDC)
        return(0)

    for full_filename in os.listdir(program_input_path):  # use the directory name here
        filename, file_ext = os.path.splitext(full_filename)
        if file_ext == '.xlsx':
            workbook = load_workbook(filename=program_input_path + full_filename)
            first_sheet = workbook.active
            tab_filename = filename + '.txt'
            print('Converting file ' + Bcolors.HEADER + filename +
                  '.xlsx' + Bcolors.ENDC + ' to Tab delimited ' +
                  Bcolors.OKGREEN + filename + '.txt' + Bcolors.ENDC)
            try:
                tab_text = open(program_input_path + tab_filename, 'w')
            except Exception as e:
                errno, strerror = e.args
                print(Bcolors.FAIL + f"I/O error({errno}): {strerror} {program_input_path}{tab_text}" + Bcolors.ENDC)
                return(0)
            else:
                txt_writer = csv.writer(tab_text, delimiter='\t', lineterminator='\n')  # writefile
                for row in first_sheet.iter_rows():
                    my_row = []
                    for cell in row:
                        value = cell.internal_value
                        result = isinstance(value, str)
                        if result:
                            if '\n' in value:
                                value = value.replace('\n', ' ')
                        result = isinstance(value, float)
                        if result:
                            value = int(value)
                        my_row.append(value)
                    txt_writer.writerow(my_row)  # write the lines to file`
                tab_text.close()
            try:
                file_hndl.write(tab_filename)
                file_hndl.write('\n')
            except IOError:
                print(Bcolors.FAIL + "Cannot write to file " + program_input_path + " filenames.txt" + Bcolors.ENDC)
                return(0)
    file_hndl.close()

def getSmartsheetMain(program_input_path, user_credentials):
    """ main function """
    colorama.init()
    filename = program_input_path + "smartsheetGetIDs.txt"
    smartsheet_ids = get_file_info(filename)
    if not smartsheet_ids:
        return(0)
    if len(smartsheet_ids) < 2:
        print(Bcolors.FAIL + "Not enough entries in smartsheetGetIDs.txt , should be 2 entries" + Bcolors.ENDC)
        return(0)
    ss_client = smartsheet.Smartsheet(user_credentials[0])
    for ids in smartsheet_ids:
        try:
            ss_client.Reports.get_report_as_excel(ids, program_input_path)
        except Exception as e:
            print(f"You Token is not valid in credentials.txt {e}")
            print("Delete credentials.txt and restart program")
            return(0)
    convert_downloaded_files_to_tab_delimited(program_input_path)
    print("All Smartsheets Downloaded")
    return(1)
